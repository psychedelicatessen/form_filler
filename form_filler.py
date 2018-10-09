import openpyxl

# these functions fill the spreadsheet 

#specify the spreadsheet being worked on
wb = openpyxl.load_workbook('type1.xlsx')

# define the number of columns in a 'fields' array for easy iterating
fields = [1,2,3,4,5,6,7]


# redefine fields for use with transport
transfields = [2,3,4,5,6,7,8]

# define a reusable function to fill the rows of the spreadsheet
def fill_row(assignment, position, date, dateNum):

    # specify the sheet the data is being added to
    sheet = wb['Payment Form']

    for i in fields:
        if i == 1:
            sheet.cell(row=position, column=i).value = assignment['Name']
        elif i == 2:
            sheet.cell(row=position, column=i).value = assignment['Activity']
        elif i == 3:
            sheet.cell(row=position, column=i).value = assignment['Time']
        elif i == 4:
            sheet.cell(row=position, column=i).value = assignment['Hours']
        elif i == 5:
            sheet.cell(row=position, column=i).value = date[dateNum]
        elif i == 6:
            sheet.cell(row=position, column=i).value = assignment['Hours']
        else:
            sheet.cell(row=position, column=i).value = assignment['Rate']
    
    return sheet
# refactored function to fill the rows of the transport sheet
def fill_row_transport(transport):
    for i in transfields:
        if i == 2:
            sheet.cell(row=position, column=i).value = transport['Name']
        elif i == 3:
            sheet.cell(row=position, column=i).value = transport['Location']
        elif i == 4:
            sheet.cell(row=position, column=i).value = transport['Type']
        elif i == 5:
            sheet.cell(row=position, column=i).value = date[dateNum]
        elif i == 7:
            sheet.cell(row=position, column=i).value = transport['Time']
        elif i == 8:
            sheet.cell(row=position, column=i).value = transport['Rate']
