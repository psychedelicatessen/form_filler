import openpyxl
import pickle
from create_assignments import *
from assignment_selector import *
# select the excel spreadsheet for editing
wb = openpyxl.load_workbook('type1.xlsx')


# define the number of columns in a 'fields' array for easy iterating
fields = [1,2,3,4,5,6,7]

# define a reusable function to fill the rows of the spreadsheet
def fill_row(assignment):
    for i in fields:
        if i == 1:
            sheet.cell(row=position, column=i).value = assignment['Name']
        elif i == 2:
            sheet.cell(row=position, column=i).value = assignment['Activity']
        elif i == 3:
            sheet.cell(row=position, column=i).value = assignment['Time']
        elif i == 4:
            sheet.cell(row=position, column=i).value = assignment['Hours']
        elif i == 5:
            sheet.cell(row=position, column=i).value = date[dateNum]
        elif i == 6:
            sheet.cell(row=position, column=i).value = assignment['Hours']
        else:
            sheet.cell(row=position, column=i).value = assignment['Rate']


# redefine fields for use with the transportation sheet
transfields = [2,3,4,5,6,7,8]

# refactored function to fill the rows of the transport sheet
def fill_row_transport(transport):
    for i in transfields:
        if i == 2:
            sheet.cell(row=position, column=i).value = transport['Name']
        elif i == 3:
            sheet.cell(row=position, column=i).value = transport['Location']
        elif i == 4:
            sheet.cell(row=position, column=i).value = transport['Type']
        elif i == 5:
            sheet.cell(row=position, column=i).value = date[dateNum]
        elif i == 7:
            sheet.cell(row=position, column=i).value = transport['Time']
        elif i == 8:
            sheet.cell(row=position, column=i).value = transport['Rate']

# variables to be filled
# initialize the variables to be filled by create functions
assignments = []
transports = []

# define the commandline prompt
prompt = ">"

# a function which grabs the date range for the assignment
def grab_date():
    print('What date range is this entry for? List the date separated by commas, no spaces. MM-DD,MM-DD,etc.')
    user_input = input(prompt)
    date = list(user_input.split(','))
    
    return date

# user's name
name = None

# user input's their name
if name == None:
    print('Please enter your name.')
    name = input(prompt)

# grab the pay period
print(f"\nHello, {name}. Please enter the Pay Period (MM-DD MM-DD i.e. 07-01 07-16)")

period = input(prompt)

# grab the submission date of the payment form
print(f"\n{period}? Okay. Now enter the submission date.")

subdate = input(prompt)

while True:
        print(f'''
        ┌─────────────────────────┐
        │         Welcome         │
        │ Please select an option │
        ╞═════════════════════════╡
        │                         │
        │(C)reate New Entry       │
        │(E)dit Data              │
        │(F)ill Payform           │
        │(L)oad Data              │
        │E(X)it & Save            │
        │                         │
        └═════════════════════════┘
        ''')

        mode = None
        choice = input(prompt)
        if choice == 'C':
            print('Create (A)ssignment or (T)ransport?')
            mode = input(prompt)
            if mode == "A":
                count = 1
                while count < 10:
                    print("Create a new assignment?")
                    answer = input(prompt)
                    if answer == 'Y':
                        assignments.append(create_assignment())
                    elif answer == 'N':
                        break
                    else:
                        print('Please enter Y or N')
            elif mode == 'T':
                count = 1
                while count < 10:
                    print("Create new transport?")
                    answer = input(prompt)
                    if answer == 'Y':
                        transports.append(create_transport())
                    elif answer == 'N':
                        break
                    else:
                        print('Please enter Y or N')
            
            else:
                
                print('Please enter either A or T')

        # view/edit the information collected thus far
        elif choice == 'E':
            print('Editing (A)ssignments or (T)ransports?')
            mode = input(prompt)

            if mode == 'A':
                assign_select(assignments)
            elif mode == 'T':
                trans_select(transports)
            else:
                print("Please enter A or T")

        # the following logic fills the spreadsheet with the collected
        # assignment/transport information
        elif choice == 'F':
            if assignments == [] or transports == []:
                print('There is no data. Please create an assignment & transport to continue.')
                pass

            else:

                print('Add an (A)ssignment or a (T)ransport?')
                mode = input(prompt)
                if mode == 'A':
                     
                    # grab the Payment Form sheet for editing
                    sheet = wb['Payment Form']

                    # input the name, period, and date into the sheet
                    sheet['A5'] = 'Instructor: ' + name
                    sheet['D5'] = 'Period: ' + period
                    sheet['F5'] = 'Submitted: ' + subdate
                    
                    # set the spreadsheet row position with an integer
                    position = 8
                
                    # select and add the specified assignment
                    # to the spreadsheet
                    while True:
                        answer == 'Y'

                        target = assign_select(assignments)

                        date = grab_date()

                        dateNum = 0

                        if len(date) > 1:

                            while dateNum < len(date):

                                fill_row(target)

                                position += 1

                                dateNum += 1
                        else:

                            fill_row(target)

                            position += 1

                        print('Your assignment was successfully added.')
                        print('Add another?')
                        answer = input(prompt)

                        if answer == 'N':

                            break
                elif mode == 'T':
                    # grab the transportation sheet for editing
                    sheet = wb['Transportation']

                    # set the name, period, and submission date
                    sheet['B5'] = 'Instructor: ' + name
                    sheet['D5'] = 'Period: ' + period
                    sheet['G5'] = 'Submitted: ' + subdate

                    # reset position
                    position = 8
                
                    while True:
                        answer == 'Y'

                        target = trans_select(transports)

                        date = grab_date()

                        dateNum = 0

                        if len(date) > 1:

                            while dateNum < len(date):

                                fill_row_transport(target)

                                position += 1

                                dateNum += 1
                        else:

                            fill_row_transport(target)

                            position += 1

                        print('Did you buy a drink?')

                        drink_answer = input(prompt)
                        
                        if drink_answer == 'Y':
                            print('How much did you pay?')
                            amount = input(prompt)

                            sheet.cell(row=position, column=2).value = target['Name']
                            sheet.cell(row=position, column=3).value = 'Drink'
                            sheet.cell(row=position, column=5).value = date[dateNum]
                            sheet.cell(row=position, column=7).value = target['Time']
                            sheet.cell(row=position, column=8).value = amount
                        
                            position += 1
                    
                        print('Your transport was successfully added.')
                        print('Add another?')
                        answer = input(prompt)

                        if answer == 'N':

                            break
                else:
                    print('Please enter either A or T.')
        elif choice == 'L':
            try:
                with open('assignment_data.txt', 'rb') as assignment_restore, open('transport_data.txt', 'rb') as transport_restore:
                    assignments = pickle.load(assignment_restore)
                    transports = pickle.load(transport_restore)
            except IOError as err:

                print('File error: ' + str(err))
            print('Load successful.')
        elif choice == 'X':
            if assignments == [] or transports == []:

                pass

            else:
                try:
                    with open('assignment_data.txt', 'wb') as assignment_save, open('transport_data.txt', 'wb') as transport_save:
                        pickle.dump(assignments, assignment_save)
                        pickle.dump(transports, transport_save)
                except IOError as err:

                    print('File error: ' + str(err))

                except pickle.PickleError as perr:

                    print('Pickling error: ' + str(perr))
                
                # format a string for the filename
                name = name.replace(" ", "")
                period = period.replace(" ", "")

                title = name.strip() + period.strip() + '.xlsx'

                print('Now saving as: ' + title + ' Goodbye.')
                wb.save(title)
            break
        
        else:
            
            print('You entered an invalid option.')
