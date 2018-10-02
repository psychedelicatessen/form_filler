################################################
# Author: https://github.com/psychedelicatessen#
################################################

import openpyxl
import pickle

# define the commandline prompt
prompt = ">"


# define a function which creates a self contained assignment object.
# this allows for the assignment to be saved for future use across payform creations and is callable for use multiple times.
def create_assignment():

    assignment = dict()

    print("What name shall I give the assignment?")
    assignment['Name'] = input(prompt)

    print("What is the designation? (I/O/P/T)")
    assignment['Activity'] = input(prompt)

    print("What time did the assignment take place? (HH:MM - HH:MM)")
    assignment['Time'] = input(prompt)

    print("Total number of billable hours spent at the assignment.")
    assignment['Hours'] = input(prompt)

    print("What is your rate for the assignment?")
    assignment['Rate'] = input(prompt)
    
    return assignment

def create_transport():

    transport = dict()
    print('Input the name of the assignment.')
    transport['Name'] = input(prompt)

    print("Please input the assignment's location.")
    transport['Location'] = input(prompt)

    print("Please input the transportation type. (TF/TX)")
    transport['Type'] = input(prompt)

    print("Please add the number of visits.")
    transport['Time'] = input(prompt)

    print("Please add the cost of travel.")
    transport['Rate'] = input(prompt)
    
    return transport


# define a reusable function to fill the rows of the spreadsheet
def fill_row(assignment):
    for i in fields:
        if i == 1:
            sheet.cell(row=position, column=i).value = assignment['Name']
        elif i == 2:
            sheet.cell(row=position, column=i).value = assignment['Activity']
        elif i == 3:
            sheet.cell(row=position, column=i).value = assignment['Time']
        elif i == 4:
            sheet.cell(row=position, column=i).value = assignment['Hours']
        elif i == 5:
            sheet.cell(row=position, column=i).value = date[dateNum]
        elif i == 6:
            sheet.cell(row=position, column=i).value = assignment['Hours']
        else:
            sheet.cell(row=position, column=i).value = assignment['Rate']

# refactored function to fill the rows of the transport sheet
def fill_row_transport(transport):
    for i in transfields:
        if i == 2:
            sheet.cell(row=position, column=i).value = transport['Name']
        elif i == 3:
            sheet.cell(row=position, column=i).value = transport['Location']
        elif i == 4:
            sheet.cell(row=position, column=i).value = transport['Type']
        elif i == 5:
            sheet.cell(row=position, column=i).value = date[dateNum]
        elif i == 7:
            sheet.cell(row=position, column=i).value = transport['Time']
        elif i == 8:
            sheet.cell(row=position, column=i).value = transport['Rate']


# a function for user assignment choice
def assign_select():
    
    count = 0

    print('Choose an assignment to add to the spreadsheet.')
    for i in trackAssignments:
        if i == 1 and assignment1 != None:
            print('1. ' + assignment1['Name'])
            count += 1
        elif i == 2 and assignment2 != None:
            print('2. ' + assignment2['Name'])
            count += 1
        elif i == 3 and assignment3 != None:
            print('3. ' + assignment3['Name'])
            count += 1
        elif i == 4 and assignment4 != None:
            print('4. ' + assignment4['Name'])
            count += 1
        elif i == 5 and assignment5 != None:
            print('5. ' + assignment5['Name'])
            count += 1

    # a loop which gets the user's target assignment for input
    while True:
        try:
            target = int(input(prompt))
        except ValueError:
            print('Enter a number between 1 and ' + str(count))
            continue
        if target > count:
            print('Your choice can\'t be greater than ' + str(count) +'!')
        else:
            break

    # set the target as the chosen assignment dictionary
    if target == 1:
        target = assignment1
    elif target == 2:
        target = assignment2
    elif target == 3:
        target = assignment3
    elif target == 4:
        target = assignment4
    elif target == 5:
        target = assignment5

    return target

# a function for user transport choice
def trans_select():
    
    count = 0

    print('Choose transport to add to the spreadsheet.')
    for i in trackAssignments:
        if i == 1 and transport1 != None:
            print('1. ' + transport1['Name'])
            count += 1
        elif i == 2 and transport2 != None:
            print('2. ' + transport2['Name'])
            count += 1
        elif i == 3 and transport3 != None:
            print('3. ' + transport3['Name'])
            count += 1
        elif i == 4 and transport4 != None:
            print('4. ' + transport4['Name'])
            count += 1
        elif i == 5 and transport5 != None:
            print('5. ' + transport5['Name'])
            count += 1

    # a loop which gets the user's target assignment for input
    while True:
        try:
            target = int(input(prompt))
        except ValueError:
            print('Enter a number between 1 and ' + str(count))
            continue
        if target > count:
            print('Your choice can\'t be greater than ' + str(count) +'!')
        else:
            break

    # set the target as the chosen transport dictionary
    if target == 1:
        target = transport1
    elif target == 2:
        target = transport2
    elif target == 3:
        target = transport3
    elif target == 4:
        target = transport4
    elif target == 5:
        target = transport5

    return target


def grab_date():
    print('What date range is this entry for? List the date separated by commas, no spaces. MM-DD,MM-DD,etc.')
    user_input = input(prompt)
    date = list(user_input.split(','))
    
    return date

# initialize the variables to be filled by create functions
trackAssignments = [1,2,3,4,5]
assignment1 = None
assignment2 = None
assignment3 = None
assignment4 = None
assignment5 = None

transport1 = None
transport2 = None
transport3 = None
transport4 = None
transport5 = None

# logic for the following if statement assignment constructor.
assNum = 1
answer = None

# a continuous loop which iterates through the variables assigned above
# it assigns all variables except for the date which will be assigned later

while True:
    
    if answer == 'N':
        print('\nOkay, all done.')
        break


    print('Create a new assignment? (Y/N)')
    answer = input(prompt)

    if answer == 'Y' and assNum == 1:

        assignment1 = create_assignment()
        print('\nOkay, your assignment was created.')        
        assNum += 1

    elif answer == 'Y' and assNum == 2:

        assignment2 = create_assignment()
        print('\nOkay, your assignment was created.')        
        assNum += 1

    elif answer == 'Y' and assNum == 3:

        assignment3 = create_assignment()
        print('\nOkay, your assignment was created.')        
        assNum += 1

    elif answer == 'Y' and assNum == 4:

        assignment4 = create_assignment()
        print('\nOkay, your assignment was created.')        
        assNum += 1

    elif answer == 'Y' and assNum == 5:

        assignment5 = create_assignment()
        print('\nOkay, your assignment was created.')        
        assNum += 1
        
    elif answer == 'Y' and assNum > 5:

        print('\nSorry, no more space for further assignments.')
        answer = 'N'

    else:

        print('Y or N only, please.')

print('\nOkay Let\'s add the transportation now.')
# reuse assNum
assNum = 1
# reset answer
answer = 'Y'

while True:
    
    if answer == 'N':
        print('\nOkay, all done.')
        break


    print('Create new transportation? (Y/N)')
    answer = input(prompt)

    if answer == 'Y' and assNum == 1:

        transport1 = create_transport()
        print('\nOkay, your transport was created.')        
        assNum += 1

    elif answer == 'Y' and assNum == 2:

        transport2 = create_transport()
        print('\nOkay, your transport was created.')        
        assNum += 1

    elif answer == 'Y' and assNum == 3:

        transport3 = create_transport()
        print('\nOkay, your transport was created.')        
        assNum += 1

    elif answer == 'Y' and assNum == 4:

        transport4 = create_transport()
        print('\nOkay, your transport was created.')        
        assNum += 1

    elif answer == 'Y' and assNum == 5:

        transport5 = create_transport()
        print('\nOkay, your transport was created.')        
        assNum += 1
        
    elif answer == 'Y' and assNum > 5:

        print('\nSorry, no more space for further transports.')
        answer = 'N'

    else:

        print('Y or N only, please.')


# open the excel file for editing and grab the Payment sheet
wb = openpyxl.load_workbook('type1.xlsx')

sheet = wb['Payment Form']

# initialize by grabbing the user's name, the date period, and the submission date.
print("Please enter your name.")

name = input(prompt)

sheet['A5'] = 'Instructor: ' + name

print(f"\nHello, {name}. Please enter the Pay Period (MM-DD MM-DD i.e. 07-01 07-16)")

period = input(prompt)

sheet['D5'] = 'Period: ' + period

print(f"\n{period}? Okay. Now enter the submission date.")

subdate = input(prompt)

sheet['F5'] = 'Submitted: ' + subdate

# logic for writing to the excel spreadsheet
#set the row position with an integer
position = 8

# define the number of columns in a 'fields' array for easy iterating
fields = [1,2,3,4,5,6,7]

# count variable for keeping track of date
dateNum = 0

# call spreadsheet population functions
target = assign_select()
date = grab_date()


# initial data adding run

if len(date) > 1:

    while dateNum < len(date):
            
        fill_row(target)

        position += 1

        dateNum += 1

else:

    fill_row(target)

    position += 1

# ask for recursion
print('\nAdd another assignment?')
answer = input(prompt)


# if user answers yes, go into assignment adding loop
if answer == 'Y':

    while True:
        if answer == 'Y':

            # reset target, date, dateNum variables for new assignment
            dateNum = 0
            
            target = assign_select()
            
            date = grab_date()

            if len(date) > 1:

                while dateNum < len(date):

                    fill_row(target)

                    position += 1

                    dateNum += 1

            else:

                fill_row(target)

                position += 1
    
            print('\nAdd another assignment?')
            answer = input(prompt)

        elif answer == 'N':

            print('\nOkay, all done.')
            break

        else:

            print('Y or N only, please')

            answer = input(prompt)

print('\nNow let\'s fill out the transportation sheet.')

# switch to the transportation sheet for editing
sheet = wb['Transportation']

# reset position value
position = 8

# redefine fields
transfields = [2,3,4,5,6,7,8]

# set the name, period, and submission date
sheet['B5'] = 'Instructor: ' + name
sheet['D5'] = 'Period: ' + period
sheet['G5'] = 'Submitted: ' + subdate

# count variable for keeping track of date
dateNum = 0

# call spreadsheet population functions
target = trans_select()
date = grab_date()

# initial data adding run

if len(date) > 1:

    while dateNum < len(date):
            
        fill_row_transport(target)

        position += 1

        dateNum += 1

else:

    fill_row_transport(target)

    position += 1

print('\nDid you buy a drink?')
drink_answer = input(prompt)

if drink_answer == 'Y':
    print('How much did you pay?')
    amount = input(prompt)
            
    sheet.cell(row=position, column=2).value = target['Name']
    sheet.cell(row=position, column=3).value = 'Drink'
    sheet.cell(row=position, column=5).value = date[dateNum]
    sheet.cell(row=position, column=7).value = target['Time']
    sheet.cell(row=position, column=8).value = amount

    position += 1

# ask for recursion
print('\nAdd more transport?')
answer = input(prompt)


# if user answers yes, go into transport adding loop
if answer == 'Y':

    while True:
        if answer == 'Y':

            # reset target, date, dateNum variables for new assignment
            dateNum = 0

            target = trans_select()
        
            date = grab_date()

            if len(date) > 1:

                while dateNum < len(date):

                    fill_row_transport(target)

                    position += 1

                    dateNum += 1

            else:

                fill_row_transport(target)

                position += 1
    
            print('\nDid you buy a drink?')
            drink_answer = input(prompt)

            if drink_answer == 'Y':
                print('How much did you pay?')
                amount = input(prompt)
            
                sheet.cell(row=position, column=2).value = target['Name']
                sheet.cell(row=position, column=3).value = 'Drink'
                sheet.cell(row=position, column=5).value = date[dateNum]
                sheet.cell(row=position, column=7).value = target['Time']
                sheet.cell(row=position, column=8).value = amount

                position += 1


            print('\nAdd more transport?')
            answer = input(prompt)

        elif answer == 'N':

            print('\nOkay, all done.')
            break

        else:

            print('\nY or N only, please')

            answer = input(prompt)

# format a string for the filename
name = name.replace(" ", "")
period = period.replace(" ", "")

title = name.strip() + period.strip() + '.xlsx'

print('Now saving as: ' + title + ' Goodbye.')
wb.save(title)

